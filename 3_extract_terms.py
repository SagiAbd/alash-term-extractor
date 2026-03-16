#!/usr/bin/env python3
"""
Script to extract scientific terms from OCR results using Gemini API.
Reads ocr/<book_id>.json and outputs terms/<book_id>.xlsx.

Supports incremental saves: the xlsx is updated after every LLM call so you
can stop at any time and resume where you left off.
"""

import json
import logging
import os
import re
import time
import argparse
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import google.generativeai as genai
from google.generativeai.types import HarmCategory, HarmBlockThreshold
from config import (
    OUTPUT_BASE_DIR,
    TERMS_OVERLAP_CHARS as OVERLAP_CHARS,
    TERMS_MODEL_NAME as MODEL_NAME,
TERMS_FALLBACK_MODELS as FALLBACK_MODEL_NAMES,
    PARALLEL_REQUESTS,
    CONST_YEAR,
    CONST_LINK,
    CONST_AUTHOR,
    book_dir_name,
)
try:
    from config import CONST_TITLE
except ImportError:
    CONST_TITLE = ""

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


def load_dotenv(env_path: Path = Path(".env")) -> None:
    """Load KEY=VALUE pairs from .env into process env without overwriting existing vars."""
    if not env_path.exists():
        return
    try:
        for raw in env_path.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if line.startswith("export "):
                line = line[len("export "):].strip()
            if "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip()
            if not key:
                continue
            if (
                (value.startswith('"') and value.endswith('"'))
                or (value.startswith("'") and value.endswith("'"))
            ):
                value = value[1:-1]
            os.environ.setdefault(key, value)
    except OSError as exc:
        log.warning("Could not read %s: %s", env_path, exc)


def configure_genai(api_key: str | None = None):
    """Configure the Gemini API. Uses the provided key, or falls back to GEMINI_API_KEY env var."""
    key = api_key or os.getenv("GEMINI_API_KEY")
    if not key:
        raise ValueError("API key required: pass --api-key or set GEMINI_API_KEY in .env")
    genai.configure(api_key=key)


def parse_page_num(page_data: Dict[str, Any]) -> int:
    """Parse page number from OCR entry; returns -1 if missing/invalid."""
    page_raw = page_data.get("page", -1)
    try:
        return int(page_raw)
    except (TypeError, ValueError):
        return -1

def load_ocr_results(filepath: Path) -> List[Dict[str, Any]]:
    """Load OCR results from JSON file."""
    if not filepath.exists():
        log.error("Input file not found: %s", filepath)
        return []

    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        log.error("Error decoding JSON file: %s", e)
        return []

def create_extraction_prompt(text: str, prev_tail: str = "", next_head: str = "") -> str:
    """Create the prompt for term extraction."""
    context_block = ""
    if prev_tail:
        context_block += f"\n[АЛДЫҢҒЫ БЕТ СОҢЫ — тек контекст үшін]:\n\"\"\"\n{prev_tail}\n\"\"\"\n"
    context_block += f"\n[НЕГІЗГІ МЕТін — осы беттен ғана терминдер алыңыз]:\n\"\"\"\n{text}\n\"\"\""
    if next_head:
        context_block += f"\n\n[КЕЛЕСІ БЕТ БАСЫ — тек контекст үшін]:\n\"\"\"\n{next_head}\n\"\"\"\n"

    return f"""Сіз қазақ тілінің тарихы мен Алаш кезеңі (ХХ ғасырдың басы) ғылымы бойынша сарапшысыз.

**МАҚСАТ**: Мәтіннен қазақ тілінде ХХ ғасыр басында ғылым мен білімнің бар болғанын дәлелдейтін сөздерді теріп алу.

Мәтін (терминдерді тек НЕГІЗГІ МӘТІННЕН алыңыз):
{context_block}

### ТАҢДАУ КРИТЕРИЙІ

Мәтіннен **ХХ ғасыр басындағы қазақ тілінде ғылым мен білімнің бар болғанын дәлелдейтін** сөздер мен сөз тіркестерін алыңыз. Бұл — физика, химия, биология, медицина, заң, экономика, философия, педагогика, техника және т.б. салалардың кәсіби немесе ғылыми лексикасы.

**Термин мүмкіндігінше қысқа болуы тиіс** — бір сөз немесе нақты мағынасы бар ең қысқа тіркес. Егер жеке сөз жеткілікті болса, ұзын тіркес алмаңыз.

### АЛМАҢЫЗ:
- Бүгінгі күнде де жиі қолданылатын қарапайым сөздер: "мұғалім", "кітап", "адам", "үй", "бару"
- Жай әкімшілік атаулар: "облысы", "уезі", "болысы", "ауданы"
- Жалқы есімдер (адам, жер аттары)
- Күмән туса — алмаңыз.

### is_definition ЕРЕЖЕСІ
`is_definition: true` — тек мәтінде термин нақты **анықталғанда** ("X — бұл Y" немесе "X деп аталатын Y" сияқты толық түсіндірме сөйлем болса).
`is_definition: false` — термин жай қолданылған. `alash_definition` өрісін `""` қалдырыңыз.

### JSON ШЫҒАРЫЛЫМЫ (тек JSON, басқа ештеңе жоқ):
{{
  "terms": [
    {{
      "alash_term": "терминнің негізгі түрі: септік/көптік/тәуелдік жалғауларын алып тастап, атау тұлғасында жаз. Мәтіндегі емлені сақта, мағынасын өзгертпе. Мысал: 'заң қызметіне' → 'заң қызметі', 'тергеу орындарының' → 'тергеу орындары'.",
      "modern_term": "қазіргі қазақ тіліндегі баламасы",
      "field": "ғылым саласы",
      "modern_definition": "қазіргі ғылыми анықтамасы (міндетті)",
      "alash_definition": "мәтіннен дәлме-дәл анықтама (немесе \"\")",
      "is_definition": true,
      "context": "термин кездесетін сөйлем + 1-2 қоршаған сөйлем (дәлме-дәл)"
    }}
  ]
}}
"""

def extract_terms_from_page(
    page_data: Dict[str, Any],
    models: List,
    prev_text: str = "",
    next_text: str = "",
) -> Optional[List[Dict[str, Any]]]:
    """Extract terms from a single page using Gemini API.

    Tries each model in *models* in order. Within each model up to max_attempts
    retries are made for transient errors. A blocked response (PROHIBITED_CONTENT)
    immediately moves to the next fallback model.

    Returns None if the page was skipped without calling the API (e.g. empty
    text), so the caller can avoid marking it as processed.  Returns a list
    (possibly empty) when the API was actually called.
    """
    page_num = page_data.get("page", -1)
    text = page_data.get("text", "")

    if not text.strip():
        log.warning("Page %d has no text, skipping.", page_num)
        return None

    prev_tail = prev_text[-OVERLAP_CHARS:] if prev_text else ""
    next_head = next_text[:OVERLAP_CHARS] if next_text else ""
    prompt = create_extraction_prompt(text, prev_tail=prev_tail, next_head=next_head)

    safety_settings = {
        HarmCategory.HARM_CATEGORY_HATE_SPEECH: HarmBlockThreshold.BLOCK_NONE,
        HarmCategory.HARM_CATEGORY_HARASSMENT: HarmBlockThreshold.BLOCK_NONE,
        HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: HarmBlockThreshold.BLOCK_NONE,
        HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT: HarmBlockThreshold.BLOCK_NONE,
    }

    max_attempts = 3
    for model_idx, model in enumerate(models):
        model_name = getattr(model, "model_name", f"model[{model_idx}]")
        for attempt in range(1, max_attempts + 1):
            try:
                response = model.generate_content(
                    prompt,
                    generation_config={"response_mime_type": "application/json"},
                    safety_settings=safety_settings,
                )
            except Exception as e:
                log.error("Page %d [%s]: API error (attempt %d/%d): %s", page_num, model_name, attempt, max_attempts, e)
                if attempt < max_attempts:
                    if "429" in str(e):
                        wait = 60 * attempt
                        log.warning("Rate limit hit, waiting %ds...", wait)
                        time.sleep(wait)
                    else:
                        time.sleep(5)
                continue

            try:
                response_text = response.text
            except ValueError as e:
                log.warning(
                    "Page %d [%s]: Response blocked (attempt %d/%d): %s — trying next model.",
                    page_num, model_name, attempt, max_attempts, e,
                )
                break  # blocked: skip remaining attempts for this model

            try:
                result = json.loads(response_text)
            except json.JSONDecodeError:
                log.error("Page %d [%s]: Failed to parse JSON (attempt %d/%d).", page_num, model_name, attempt, max_attempts)
                if attempt < max_attempts:
                    time.sleep(5)
                continue

            terms = result.get("terms", [])
            enriched_terms = []
            for term in terms:
                enriched_terms.append({
                    "Заманауи термин": term.get("modern_term", ""),
                    "Алаш термині": term.get("alash_term", ""),
                    "Сала": term.get("field", ""),
                    "Заманауи түсініктеме": term.get("modern_definition", ""),
                    "Алаш түсініктемесі": term.get("alash_definition", ""),
                    "Анықтама бар ма": term.get("is_definition", False),
                    "Екі бет арасындағы мәтін -- контекст үшін": term.get("context", ""),
                    "Авторы": CONST_AUTHOR,
                    "Басталатын беті": page_num,
                    "Аяқталу беті": page_num,
                    "Жазылу жылы": CONST_YEAR,
                    "Сілтеме": CONST_LINK,
                })

            log.info("Page %d [%s]: Extracted %d terms.", page_num, model_name, len(enriched_terms))
            return enriched_terms

    log.error("Page %d: All models exhausted, page will not be marked as processed.", page_num)
    return []  # empty list = API was tried but failed; None = skipped (no text)


# ---------------------------------------------------------------------------
# Incremental state (resume support)
# ---------------------------------------------------------------------------

def load_state(state_file: Path) -> tuple[list, set]:
    """Load raw terms and processed page numbers from the state file."""
    if not state_file.exists():
        return [], set()
    try:
        data = json.loads(state_file.read_text(encoding="utf-8"))
        terms = data.get("terms", [])
        processed = set(data.get("processed_pages", []))
        log.info(
            "Resuming: %d terms already extracted from %d page(s).",
            len(terms), len(processed),
        )
        return terms, processed
    except Exception as e:
        log.warning("Could not load state file %s: %s", state_file, e)
        return [], set()


def save_state(state_file: Path, terms: list, processed_pages: set):
    """Persist raw terms and processed page set to JSON."""
    state_file.parent.mkdir(parents=True, exist_ok=True)
    state_file.write_text(
        json.dumps(
            {"terms": terms, "processed_pages": sorted(processed_pages)},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

COLUMNS = [
    "Заманауи термин", "Алаш термині", "Сала",
    "Заманауи түсініктеме", "Алаш түсініктемесі", "Анықтама бар ма",
    "Екі бет арасындағы мәтін -- контекст үшін", "Авторы",
    "Басталатын беті", "Аяқталу беті", "Жазылу жылы", "Сілтеме"
]


def save_xlsx(terms: List[Dict[str, Any]], output_path: Path):
    """Write terms to an Excel file with a metadata header block."""
    if not terms:
        return

    df = pd.DataFrame(terms)
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[COLUMNS]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _save_with_metadata_header(df, output_path)
    log.info("Saved %d terms to %s", len(df), output_path)


def _save_with_metadata_header(df: pd.DataFrame, output_path: Path):
    """
    Write *df* to an Excel file, preceded by a styled metadata block.

    Layout:
      Row 1  : Кітап атауы  | <title>
      Row 2  : Авторы        | <author>
      Row 3  : Жазылу жылы  | <year>
      Row 4  : Сілтеме      | <link>
      Row 5  : (blank)
      Row 6  : column headers (bold, size 11)
      Row 7+ : data
    """
    meta_rows = [
        ("Кітап атауы",  CONST_TITLE  or ""),
        ("Авторы",       CONST_AUTHOR or ""),
        ("Жазылу жылы",  str(CONST_YEAR) if CONST_YEAR else ""),
        ("Сілтеме",      CONST_LINK   or ""),
    ]

    meta_font   = Font(bold=True, size=24)
    header_font = Font(bold=True, size=11)

    wb = openpyxl.Workbook()
    ws = wb.active

    # --- Metadata rows ---
    for i, (label, value) in enumerate(meta_rows, start=1):
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=value)
        lc.font = meta_font
        vc.font = meta_font
        ws.row_dimensions[i].height = 36  # tall enough for 24pt font

    # Blank separator row (row 5 when 4 meta rows)
    blank_row  = len(meta_rows) + 1
    header_row = blank_row + 1

    # --- Column header row ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        hc = ws.cell(row=header_row, column=col_idx, value=col_name)
        hc.font = header_font

    # --- Data rows ---
    _illegal = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

    def _clean(v):
        if isinstance(v, str):
            return _illegal.sub("", v)
        return v

    for r_idx, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_clean(value))

    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------

def deduplicate_terms(terms: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Remove duplicate entries by alash_term (case-insensitive). Keeps the
    first occurrence (lowest page number, since the list is pre-sorted)."""
    seen: set = set()
    result = []
    for t in terms:
        key = t.get("Алаш термині", "").strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(t)
    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Extract terms from OCR results to Excel")
    parser.add_argument("--limit", type=int, help="Limit number of pages to process (for testing, index-based mode)")
    parser.add_argument("--start", type=int, default=0, help="Start page index (0-based, index-based mode)")
    parser.add_argument("--start-page", "-s", type=int, default=None, help="Start page number (inclusive)")
    parser.add_argument("--end-page", "-e", type=int, default=None, help="End page number (inclusive)")
    parser.add_argument("--pages", "-p", type=str, default=None, help="Comma-separated page numbers to process, e.g. 1,5,10,15")
    parser.add_argument("--rerun-failed", action="store_true", help="Re-process pages listed in failed_term_pages in .metadata.json")
    parser.add_argument("--api-key", type=str, default=None, help="Gemini API key (overrides GEMINI_API_KEY env var)")
    args = parser.parse_args()

    load_dotenv()

    # --- Derive file paths from author + title (set by 0_metadata_scrape.py) ---
    book_dir   = Path(OUTPUT_BASE_DIR) / book_dir_name()
    input_file  = book_dir / "ocr.json"
    output_xlsx = book_dir / "terms.xlsx"
    state_file  = book_dir / "terms_state.json"

    log.info("Book dir  : %s", book_dir)
    log.info("OCR input : %s", input_file)
    log.info("Excel out : %s", output_xlsx)

    # --- Argument validation ---
    if args.start < 0:
        parser.error("--start must be >= 0")
    if args.limit is not None and args.limit < 1:
        parser.error("--limit must be >= 1")
    if args.start_page is not None and args.start_page < 1:
        parser.error("--start-page must be >= 1")
    if args.end_page is not None and args.end_page < 1:
        parser.error("--end-page must be >= 1")
    if (
        args.start_page is not None
        and args.end_page is not None
        and args.start_page > args.end_page
    ):
        parser.error("--start-page cannot be greater than --end-page")
    if (
        (args.start_page is not None or args.end_page is not None)
        and (args.start != 0 or args.limit is not None)
    ):
        parser.error("Use either --start/--limit (index mode) or --start-page/--end-page (page mode), not both")
    if args.pages is not None and (args.start_page is not None or args.end_page is not None or args.start != 0 or args.limit is not None):
        parser.error("--pages cannot be combined with --start, --limit, --start-page, or --end-page")
    if args.rerun_failed and (args.pages is not None or args.start_page is not None or args.end_page is not None or args.start != 0 or args.limit is not None):
        parser.error("--rerun-failed cannot be combined with other page selection flags")

    explicit_pages: set[int] | None = None
    if args.rerun_failed:
        _meta_file = Path(__file__).parent / ".metadata.json"
        try:
            meta = json.loads(_meta_file.read_text(encoding="utf-8")) if _meta_file.exists() else {}
            failed = meta.get("failed_term_pages", [])
        except Exception as e:
            log.error("Could not read .metadata.json: %s", e)
            failed = []
        if not failed:
            log.info("No failed pages found in .metadata.json. Nothing to rerun.")
            return
        explicit_pages = set(failed)
        log.info("Rerunning %d failed page(s) from .metadata.json: %s", len(explicit_pages), sorted(explicit_pages))
    elif args.pages is not None:
        try:
            explicit_pages = {int(p.strip()) for p in args.pages.split(",") if p.strip()}
        except ValueError:
            parser.error("--pages must be comma-separated integers, e.g. 1,5,10,15")

    configure_genai(args.api_key)
    models = [genai.GenerativeModel(name) for name in [MODEL_NAME] + FALLBACK_MODEL_NAMES]
    log.info("Model chain: %s", " → ".join([MODEL_NAME] + FALLBACK_MODEL_NAMES))

    ocr_data = load_ocr_results(input_file)
    if not ocr_data:
        log.error("No OCR data loaded.")
        return

    # Build page-number → text index for prev/next context lookup.
    # Only adjacent pages (page_num ± 1) are used; gaps from failed OCR pages
    # will naturally produce an empty string.
    page_text: Dict[int, str] = {
        parse_page_num(p): p.get("text", "")
        for p in ocr_data
        if parse_page_num(p) >= 0
    }

    # --- Build index of pages to process ---
    selected_indices: List[int] = []
    if explicit_pages is not None:
        for i, page_data in enumerate(ocr_data):
            page_num = parse_page_num(page_data)
            if page_num in explicit_pages:
                selected_indices.append(i)
        matched = {parse_page_num(ocr_data[i]) for i in selected_indices}
        missing = explicit_pages - matched
        if missing:
            log.warning("Pages not found in OCR data: %s", sorted(missing))
        log.info("Processing %d specific page(s): %s", len(selected_indices), sorted(matched))
    elif args.start_page is not None or args.end_page is not None:
        skipped_invalid = 0
        for i, page_data in enumerate(ocr_data):
            page_num = parse_page_num(page_data)
            if page_num < 1:
                skipped_invalid += 1
                continue
            if args.start_page is not None and page_num < args.start_page:
                continue
            if args.end_page is not None and page_num > args.end_page:
                continue
            selected_indices.append(i)
        if skipped_invalid:
            log.warning("Skipped %d OCR record(s) with invalid page numbers.", skipped_invalid)
        if not selected_indices:
            log.warning(
                "No OCR records matched page range %s-%s.",
                args.start_page if args.start_page is not None else "*",
                args.end_page if args.end_page is not None else "*",
            )
            return
        log.info(
            "Processing %d OCR record(s) in page range %s-%s.",
            len(selected_indices),
            args.start_page if args.start_page is not None else "*",
            args.end_page if args.end_page is not None else "*",
        )
    else:
        start_idx = args.start
        end_idx = start_idx + args.limit if args.limit else len(ocr_data)
        if start_idx >= len(ocr_data):
            log.warning(
                "Start index %d is out of range for OCR dataset of size %d.",
                start_idx,
                len(ocr_data),
            )
            return
        end_idx = min(end_idx, len(ocr_data))
        selected_indices = list(range(start_idx, end_idx))
        log.info("Processing OCR indices %d to %d.", start_idx, end_idx - 1)

    # --- Load existing state for resume ---
    all_terms, processed_pages = load_state(state_file)

    # --- Main extraction loop (parallel) ---
    pending_indices = [
        i for i in selected_indices
        if parse_page_num(ocr_data[i]) not in processed_pages
    ]
    skipped = len(selected_indices) - len(pending_indices)
    if skipped:
        log.info("%d page(s) already processed, skipping.", skipped)

    state_lock = threading.Lock()
    processed_count = 0
    dupes = 0
    failed_pages: list[int] = []

    def process_page(src_idx: int):
        page = ocr_data[src_idx]
        page_num = parse_page_num(page)
        prev_text = page_text.get(page_num - 1, "")
        next_text = page_text.get(page_num + 1, "")
        return page_num, extract_terms_from_page(page, models, prev_text=prev_text, next_text=next_text)

    with ThreadPoolExecutor(max_workers=PARALLEL_REQUESTS) as executor:
        futures = [executor.submit(process_page, idx) for idx in pending_indices]
        for future in as_completed(futures):
            page_num, terms = future.result()
            if terms is None:
                # Skipped naturally (empty text) — leave unprocessed for retry
                continue
            with state_lock:
                if len(terms) == 0 and page_num >= 0:
                    # API was attempted but all models failed
                    if page_num not in failed_pages:
                        failed_pages.append(page_num)
                    log.warning("Page %d marked as failed (all models exhausted).", page_num)
                else:
                    all_terms.extend(terms)
                    all_terms.sort(key=lambda t: t.get("Басталатын беті", -1))
                    before_dedup = len(all_terms)
                    all_terms = deduplicate_terms(all_terms)
                    dupes = before_dedup - len(all_terms)
                    if page_num >= 0:
                        processed_pages.add(page_num)
                    save_state(state_file, all_terms, processed_pages)
                    save_xlsx(all_terms, output_xlsx)
                    processed_count += 1
                    if processed_count % 10 == 0:
                        log.info("Processed %d new page(s)...", processed_count)

    _meta_file = Path(__file__).parent / ".metadata.json"
    if failed_pages or args.rerun_failed:
        failed_pages.sort()
        if failed_pages:
            log.warning("Failed pages (all models exhausted): %s", failed_pages)
        try:
            meta = json.loads(_meta_file.read_text(encoding="utf-8")) if _meta_file.exists() else {}
            if args.rerun_failed:
                # Remove pages that succeeded this run; keep ones still failing
                still_failing = set(failed_pages)
                meta["failed_term_pages"] = sorted(still_failing)
                if not still_failing:
                    log.info("All previously failed pages succeeded. Cleared failed_term_pages in .metadata.json.")
                else:
                    log.warning("Still failing after rerun: %s", sorted(still_failing))
            else:
                existing = set(meta.get("failed_term_pages", []))
                existing.update(failed_pages)
                meta["failed_term_pages"] = sorted(existing)
            _meta_file.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
            if failed_pages:
                log.info("Failed pages written to .metadata.json: %s", meta["failed_term_pages"])
        except Exception as e:
            log.error("Could not update .metadata.json with failed pages: %s", e)

    if processed_count == 0:
        log.info("No new pages to process.")
    else:
        log.info(
            "Done. Processed %d new page(s). Unique terms: %d. Duplicates removed: %d.",
            processed_count, len(all_terms), dupes,
        )


if __name__ == "__main__":
    _t0 = time.time()
    main()
    _elapsed = time.time() - _t0
    log.info("Total time: %dm %02ds", int(_elapsed // 60), int(_elapsed % 60))
    if os.environ.get("PLAY_SOUND", "").lower() in ("1", "true", "yes"):
        _snd = Path(__file__).parent / "done.mp3"
        if _snd.exists():
            os.system(f'afplay -t 10 "{_snd}"')
